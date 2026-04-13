"""
Generate IFRS17 unit test Panel rows from the Model Points_Profiles planned test list.

Reads all 3480 planned test names from the Profiles sheet, assigns meaningful
parameter values based on each test's 6 dimensions, then writes:
  1. A CSV for review: backend/unit_tests_panel.csv
  2. A copy of the xlsm with the Panel populated: backend/IFRS17_UnitTests_v1.xlsm

Run from the repo root:
    C:/Python314/python.exe backend/generate_panel.py
"""

import csv
import shutil
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSM = REPO_ROOT / ".claude" / "IFRS17 Units Testing_v16_IM.xlsm"
BACKEND = REPO_ROOT / "backend"
OUTPUT_XLSM = BACKEND / "IFRS17_UnitTests_v1.xlsm"
OUTPUT_CSV = BACKEND / "unit_tests_panel.csv"

# ---------------------------------------------------------------------------
# Model-specific base rates
# ---------------------------------------------------------------------------
BASE_PARAMS = {
    "GMM": {
        "lapse": 0.10, "benefits": 0.60, "ndic": 0.90,
        "expenses": 0.15, "commissions": 0.03, "acq_cost": 0.02,
    },
    # VFA: variable/unit-linked — lower pure insurance content
    "VFA": {
        "lapse": 0.05, "benefits": 0.40, "ndic": 0.60,
        "expenses": 0.10, "commissions": 0.02, "acq_cost": 0.01,
    },
    # PAA: short-term — higher loss ratio, smaller opening balances
    "PAA": {
        "lapse": 0.05, "benefits": 0.70, "ndic": 0.20,
        "expenses": 0.12, "commissions": 0.03, "acq_cost": 0.02,
    },
}

# ---------------------------------------------------------------------------
# Opening balances by model and opening state
# csm   = CSM opening (0 when onerous)
# lc    = LC opening  (0 when profitable)
# iacfs = Insurance acquisition cash flows asset
# nb    = new business flag (1 = initial recognition period)
# ---------------------------------------------------------------------------
OPENINGS = {
    "GMM": {
        "PROF": {"csm": 500, "lc": 0,   "iacfs": 100, "nb": 0},
        "ONER": {"csm": 0,   "lc": 500, "iacfs": 100, "nb": 0},
        "MARG": {"csm": 50,  "lc": 0,   "iacfs": 10,  "nb": 0},
        "NEWB": {"csm": 0,   "lc": 0,   "iacfs": 0,   "nb": 1},
    },
    "VFA": {
        "PROF": {"csm": 500, "lc": 0,   "iacfs": 100, "nb": 0},
        "ONER": {"csm": 0,   "lc": 500, "iacfs": 100, "nb": 0},
        "MARG": {"csm": 50,  "lc": 0,   "iacfs": 10,  "nb": 0},
        "NEWB": {"csm": 0,   "lc": 0,   "iacfs": 0,   "nb": 1},
    },
    "PAA": {
        "PROF": {"csm": 300, "lc": 0,   "iacfs": 50,  "nb": 0},
        "ONER": {"csm": 0,   "lc": 200, "iacfs": 50,  "nb": 0},
        "MARG": {"csm": 30,  "lc": 0,   "iacfs": 10,  "nb": 0},
        "NEWB": {"csm": 0,   "lc": 0,   "iacfs": 0,   "nb": 1},
    },
}

# ---------------------------------------------------------------------------
# Stress magnitudes
# ---------------------------------------------------------------------------
# Actuals-vs-expected variances (EXP_VAR, MODIF, COMB)
ACT_STR = {"STB": 0.02, "MODER": 0.15, "EXTR": 0.50}
# EoP assumption changes (ECON, DEMO, OPER, Population, MODIF, COMB)
EOP_STR = {"STB": 0.05, "MODER": 0.20, "EXTR": 0.60}
# Yield-curve parallel shifts (ECON, COMB)
RATE_STR = {"STB": 0.005, "MODER": 0.025, "EXTR": 0.100}

BASE_RATE = 0.03      # baseline current yield curve level
BASE_LOCKED = 0.03   # baseline locked-in rate (used for OCI contracts)
MIN_RATE = 0.001     # floor for yield curves (negative rates are unrealistic for this portfolio)


# ---------------------------------------------------------------------------
# Stress mapping: (driver, direction) → parameter increments
# ---------------------------------------------------------------------------
def _r(x, d=4):
    """Round to avoid floating-point noise."""
    return round(x, d)


def get_stresses(driver: str, direction: str, impact: str) -> dict:
    a = ACT_STR[impact]
    e = EOP_STR[impact]
    r = RATE_STR[impact]

    # Defaults — all zero
    eop_lapse = eop_ben = eop_ndic = eop_exp = eop_comm = eop_acq = 0.0
    act_prm = act_clm_cs = act_ndic_cs = act_exp = act_comm = act_acq = act_clm_ps = 0.0
    eop_cur = BASE_RATE   # EoP current yield (only changes for ECON/COMB)

    def _yc(val):
        """Apply floor to yield curve values — negative rates unsupported."""
        return max(val, MIN_RATE)

    # -----------------------------------------------------------------------
    if driver == "NORM":
        # Intended as pure baseline; small residuals for non-FLAT directions
        # so every row has a unique set of parameters.
        if   direction == "FLAT":        pass
        elif direction == "UP_prms":     act_prm     = _r(a * 0.2)
        elif direction == "Down_prm":    act_prm     = _r(-a * 0.2)
        elif direction == "UP_claims":   act_clm_cs  = _r(a * 0.2)
        elif direction == "Down_claims": act_clm_cs  = _r(-a * 0.2)
        elif direction == "Lapse_UP":    eop_lapse   = _r(e * 0.2)
        elif direction == "Lapse_DOWN":  eop_lapse   = _r(-e * 0.2)

    # -----------------------------------------------------------------------
    elif driver == "EXP_VAR":
        # Actual cash flows differ from expected (current period)
        if   direction == "UP_prms":
            act_prm     = a                          # premiums higher than expected
        elif direction == "Down_prm":
            act_prm     = -a                         # premiums lower than expected
        elif direction == "UP_claims":
            act_clm_cs  = a                          # more claims than expected
            act_ndic_cs = _r(a * 0.9)               # NDIC moves with claims
        elif direction == "Down_claims":
            act_clm_cs  = -a
            act_ndic_cs = _r(-a * 0.9)
        elif direction == "Lapse_UP":
            # More lapses → fewer premiums received, fewer claims paid
            act_prm     = _r(-a * 0.5)
            act_clm_cs  = _r(-a * 0.3)
        elif direction == "Lapse_DOWN":
            act_prm     = _r(a * 0.5)
            act_clm_cs  = _r(a * 0.3)
        elif direction == "FLAT":
            pass                                     # zero variance — as expected

    # -----------------------------------------------------------------------
    elif driver == "ECON":
        # Changes in economic variables affecting the EoP yield curve
        # UP_prms / UP_claims  → rates rise (asset values fall, discount rate rises)
        # Down_prm / Down_claims → rates fall
        if   direction in ("UP_prms", "UP_claims"):
            eop_cur = _r(BASE_RATE + r)
        elif direction in ("Down_prm", "Down_claims"):
            eop_cur = _yc(_r(BASE_RATE - r))
        elif direction == "Lapse_UP":
            # Rate shock also drives lapse behaviour (surrender at rate rise)
            eop_cur   = _r(BASE_RATE + r)
            eop_lapse = _r(e * 0.5)
        elif direction == "Lapse_DOWN":
            eop_cur   = _yc(_r(BASE_RATE - r))
            eop_lapse = _r(-e * 0.5)
        elif direction == "FLAT":
            eop_cur = BASE_RATE                      # no economic movement

    # -----------------------------------------------------------------------
    elif driver == "DEMO":
        # Changes in demographic / biometric assumptions (EoP projections)
        if   direction == "UP_prms":
            # Better persistency → effective premium capacity rises
            eop_lapse = _r(-e)                       # lapse assumption improves
        elif direction == "Down_prm":
            eop_lapse = _r(e)                        # lapse assumption worsens
        elif direction == "UP_claims":
            # Worsening mortality / morbidity
            eop_ben  = e
            eop_ndic = _r(e * 0.9)
        elif direction == "Down_claims":
            # Improving mortality / morbidity
            eop_ben  = -e
            eop_ndic = _r(-e * 0.9)
        elif direction == "Lapse_UP":
            eop_lapse = _r(e)
        elif direction == "Lapse_DOWN":
            eop_lapse = _r(-e)
        elif direction == "FLAT":
            pass

    # -----------------------------------------------------------------------
    elif driver == "OPER":
        # Changes in operational / expense assumptions (EoP projections)
        if   direction == "UP_prms":
            # Operational efficiency → lower acquisition costs
            eop_acq  = _r(-e)
        elif direction == "Down_prm":
            # Operational inefficiency → higher acquisition costs
            eop_acq  = _r(e)
        elif direction == "UP_claims":
            # Higher ongoing expenses and commissions
            eop_exp  = e
            eop_comm = _r(e * 0.3)
        elif direction == "Down_claims":
            eop_exp  = -e
            eop_comm = _r(-e * 0.3)
        elif direction == "Lapse_UP":
            # Poor service drives higher expenses and more lapses
            eop_exp   = _r(e * 0.7)
            eop_lapse = _r(e * 0.4)
        elif direction == "Lapse_DOWN":
            eop_exp   = _r(-e * 0.7)
            eop_lapse = _r(-e * 0.4)
        elif direction == "FLAT":
            pass

    # -----------------------------------------------------------------------
    elif driver == "Population":
        # Changes in lapse / population mix assumptions (EoP projections)
        if   direction == "UP_prms":
            # Better retention → lower lapse assumption
            eop_lapse = _r(-e)
        elif direction == "Down_prm":
            # Worse retention → higher lapse assumption
            eop_lapse = _r(e)
        elif direction == "UP_claims":
            # Higher risk population mix → higher benefit / NDIC assumptions
            eop_ben  = _r(e * 0.5)
            eop_ndic = _r(e * 0.4)
        elif direction == "Down_claims":
            eop_ben  = _r(-e * 0.5)
            eop_ndic = _r(-e * 0.4)
        elif direction == "Lapse_UP":
            eop_lapse = _r(e)
        elif direction == "Lapse_DOWN":
            eop_lapse = _r(-e)
        elif direction == "FLAT":
            pass

    # -----------------------------------------------------------------------
    elif driver == "MODIF":
        # Contract modification — affects both current-period actuals and EoP projections
        if   direction == "UP_prms":
            # Modification that increases future premiums (coverage upgrade)
            act_prm = _r(a * 0.5)
            eop_ben = _r(e * 0.3)       # more coverage → higher projected benefits
        elif direction == "Down_prm":
            act_prm = _r(-a * 0.5)
            eop_ben = _r(-e * 0.3)
        elif direction == "UP_claims":
            # Benefit increase modification
            eop_ben = e
        elif direction == "Down_claims":
            # Benefit reduction modification
            eop_ben = -e
        elif direction == "Lapse_UP":
            # Mass modification triggering surrenders
            eop_lapse = _r(e)
            act_prm   = _r(-a * 0.3)
        elif direction == "Lapse_DOWN":
            # Modification improving retention
            eop_lapse = _r(-e)
            act_prm   = _r(a * 0.2)
        elif direction == "FLAT":
            pass                         # immaterial modification

    # -----------------------------------------------------------------------
    elif driver == "COMB":
        # Multiple simultaneous drivers — partial contributions from each
        if   direction == "UP_prms":
            # Multi-favorable: higher premiums + improving assumptions
            act_prm  = _r(a * 0.5)
            eop_ben  = _r(-e * 0.3)
            eop_exp  = _r(-e * 0.2)
        elif direction == "Down_prm":
            # Multi-adverse affecting premium side
            act_prm  = _r(-a * 0.5)
            eop_ben  = _r(e * 0.3)
            eop_exp  = _r(e * 0.2)
        elif direction == "UP_claims":
            # Multi-adverse: higher claims + expenses + rate rise
            act_clm_cs = _r(a * 0.4)
            eop_exp    = _r(e * 0.3)
            eop_cur    = _r(BASE_RATE + r * 0.5)
        elif direction == "Down_claims":
            # Multi-favorable: lower claims + expenses + rate fall
            act_clm_cs = _r(-a * 0.4)
            eop_exp    = _r(-e * 0.3)
            eop_cur    = _yc(_r(BASE_RATE - r * 0.5))
        elif direction == "Lapse_UP":
            # Adverse lapse scenario combined with higher claims
            eop_lapse  = _r(e * 0.5)
            act_clm_cs = _r(a * 0.3)
        elif direction == "Lapse_DOWN":
            # Favorable lapse scenario combined with lower claims
            eop_lapse  = _r(-e * 0.5)
            act_clm_cs = _r(-a * 0.3)
        elif direction == "FLAT":
            pass

    return {
        "eop_lapse":   eop_lapse,
        "eop_ben":     eop_ben,
        "eop_ndic":    eop_ndic,
        "eop_exp":     eop_exp,
        "eop_comm":    eop_comm,
        "eop_acq":     eop_acq,
        "act_prm":     act_prm,
        "act_clm_cs":  act_clm_cs,
        "act_ndic_cs": act_ndic_cs,
        "act_exp":     act_exp,
        "act_comm":    act_comm,
        "act_acq":     act_acq,
        "act_clm_ps":  act_clm_ps,
        "eop_cur":     eop_cur,
    }


# ---------------------------------------------------------------------------
# Assemble one Panel row (31 values matching Panel columns A–AE)
# ---------------------------------------------------------------------------
def make_panel_row(ut_name, mp_id, model, oci_opt, opening, driver, direction, impact):
    bp = BASE_PARAMS[model]
    op = OPENINGS[model][opening]
    st = get_stresses(driver, direction, impact)

    # Yield-curve columns
    bop_cur  = BASE_RATE
    eop_cur  = st["eop_cur"]

    if oci_opt == "OCI":
        # OCI elected: lock in the rate at BoP; OCI absorbs current-vs-locked gap
        bop_lock = BASE_LOCKED
        eop_lock = BASE_LOCKED   # locked-in rate does not change during the period
        oci_bal  = 0             # opening OCI cumulative balance (0 for simplicity)
    else:
        bop_lock = 0.0
        eop_lock = 0.0
        oci_bal  = 0

    return [
        ut_name,             # col A  – UT Name (6-dimension ID)
        mp_id,               # col B  – Model Point ID
        model,               # col C  – Measurement Model (GMM / VFA / PAA)
        op["nb"],            # col D  – New Business flag
        bp["lapse"],         # col E  – Base lapse rate
        bp["benefits"],      # col F  – Base benefits ratio
        bp["ndic"],          # col G  – Base NDIC ratio
        bp["expenses"],      # col H  – Base expenses ratio
        bp["commissions"],   # col I  – Base commissions ratio
        bp["acq_cost"],      # col J  – Base acquisition cost ratio
        # --- EoP assumption stresses (cols K–P) ---
        st["eop_lapse"],     # col K
        st["eop_ben"],       # col L
        st["eop_ndic"],      # col M
        st["eop_exp"],       # col N
        st["eop_comm"],      # col O
        st["eop_acq"],       # col P
        # --- Actuals vs expected (cols Q–W) ---
        st["act_prm"],       # col Q  – Premiums
        st["act_clm_cs"],    # col R  – Paid Claims – Current Service
        st["act_ndic_cs"],   # col S  – NDIC – Current Service
        st["act_exp"],       # col T  – Attributable Expenses
        st["act_comm"],      # col U  – Commissions
        st["act_acq"],       # col V  – Acquisition Costs
        st["act_clm_ps"],    # col W  – Paid Claims – Past Service
        # --- Opening balances (cols X–AA) ---
        op["csm"],           # col X  – CSM Opening Balance
        op["lc"],            # col Y  – LC Opening Balance
        op["iacfs"],         # col Z  – IACFs Opening Balance
        oci_bal,             # col AA – OCI opening balance
        # --- Yield curves (cols AB–AE) ---
        _r(bop_cur),         # col AB – BoP Current Yield Curve
        _r(bop_lock),        # col AC – BoP Locked-In Yield Curve
        _r(eop_cur),         # col AD – EoP Current Yield Curve
        _r(eop_lock),        # col AE – EoP Locked-In Yield Curve
    ]


# ---------------------------------------------------------------------------
# Read planned test list from Model Points_Profiles
# ---------------------------------------------------------------------------
print("Reading planned tests from Model Points_Profiles …")
wb_src = openpyxl.load_workbook(SOURCE_XLSM, read_only=True, keep_vba=True, data_only=True)
profiles = wb_src["Model Points_Profiles"]
planned = [
    (r[0], r[1])
    for r in profiles.iter_rows(min_row=2, values_only=True)
    if r[0] is not None
]
wb_src.close()
print(f"  {len(planned)} planned tests found.")

# ---------------------------------------------------------------------------
# Parse each test name and generate Panel rows
# ---------------------------------------------------------------------------
panel_rows = []
skipped = []

for ut_name, mp_id in planned:
    parts = ut_name.split("-")
    # Names like "GMM-No OCI-PROF-ECON-UP_prms-STB" split into exactly 6 parts
    if len(parts) != 6:
        skipped.append(ut_name)
        continue

    model, oci_opt, opening, driver, direction, impact = parts

    try:
        row = make_panel_row(ut_name, mp_id, model, oci_opt, opening, driver, direction, impact)
        panel_rows.append(row)
    except KeyError as exc:
        skipped.append(f"{ut_name} ({exc})")

print(f"  {len(panel_rows)} rows generated, {len(skipped)} skipped.")
if skipped:
    print("  Skipped:", skipped[:10])

# ---------------------------------------------------------------------------
# Write CSV
# ---------------------------------------------------------------------------
CSV_HEADERS = [
    "UT Name", "MP ID", "Model", "New Business",
    "Lapse Rate", "Benefits", "NDIC", "Expenses", "Commissions", "Acquisition Cost",
    "EoP Lapse Stress", "EoP Benefits Stress", "EoP NDIC Stress",
    "EoP Expenses Stress", "EoP Commissions Stress", "EoP Acq Cost Stress",
    "Act Premiums", "Act Paid Claims-CS", "Act NDIC-CS",
    "Act Expenses", "Act Commissions", "Act Acq Costs", "Act Paid Claims-PS",
    "CSM Opening", "LC Opening", "IACFs Opening", "OCI Opening",
    "BoP Current YC", "BoP Locked YC", "EoP Current YC", "EoP Locked YC",
]

print(f"\nWriting CSV → {OUTPUT_CSV}")
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(CSV_HEADERS)
    writer.writerows(panel_rows)
print("  Done.")

# ---------------------------------------------------------------------------
# Write populated Panel to a copy of the xlsm
# ---------------------------------------------------------------------------
print(f"\nCopying source file → {OUTPUT_XLSM}")
shutil.copy2(SOURCE_XLSM, OUTPUT_XLSM)

print("Opening copy for writing …")
wb_out = openpyxl.load_workbook(OUTPUT_XLSM, keep_vba=True, data_only=False)
ws_panel = wb_out["Panel"]

# Panel data starts at row 6 (rows 1-5 are headers / template)
PANEL_DATA_START = 6

# Clear existing data rows below row 5 before writing
print("  Clearing old data rows …")
for row_idx in range(PANEL_DATA_START, ws_panel.max_row + 1):
    for col_idx in range(1, 32):   # columns A–AE (1-indexed)
        ws_panel.cell(row=row_idx, column=col_idx).value = None

# Write generated rows
print(f"  Writing {len(panel_rows)} rows …")
for offset, row_values in enumerate(panel_rows):
    excel_row = PANEL_DATA_START + offset
    for col_idx, value in enumerate(row_values, start=1):
        ws_panel.cell(row=excel_row, column=col_idx).value = value

wb_out.save(OUTPUT_XLSM)
print(f"  Saved: {OUTPUT_XLSM}")
print("\nDone. Review the CSV first, then use the xlsm copy with your macros.")
